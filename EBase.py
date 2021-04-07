# -*- coding: utf-8 -*-
"""
Created on Sun Feb 28 11:17:39 2021

@author: Gerald
"""

# -*- coding: utf-8 -*-
import re

"""
Created on Thu Feb 25 14:44:13 2021

@author: Gerald
"""

import sqlite3

import openpyxl
from sqlalchemy import create_engine, MetaData, Table, Column, String


def read(file, first_col_id, last_col_id, data_ids, equipment):
    data_dict = {}
    time_dict = {}
    temperature_dict = {}
    read_dict = {}

    filePathsId = ("B4", "B5")
    procedureCellsId = ("B14", "B21")
    modelId = ("B9", "B9")
    experimentDateAndTimeId = ("B7", "B8")
    settingsId = ('A2', 'F2')
    identifierId = ('A2', 'A2')
    wellId = ('A1', 'CR1')
    wellInfoId = ('A2', 'CR2')
    data_sheet = 'Data'
    settings_sheet = 'Metadata'
    well_sheet = 'Well Information'

    settingsCells = get_cells(file, settingsId, settings_sheet)
    filePathsCells = get_cells(file, filePathsId, data_sheet)
    experimentDateAndTimeCells = get_cells(file, experimentDateAndTimeId, data_sheet)
    procedureCells = get_cells(file, procedureCellsId, data_sheet)
    wellCells = get_cells(file, wellId, well_sheet)
    wellInfoCells = get_cells(file, wellInfoId, well_sheet)
    model = get_single_value(modelId, file, data_sheet)
    identifier = get_identifier(identifierId, file, settings_sheet)
    equipment = create_dict(equipment, model)
    filePaths = create_file_paths_dict(filePathsCells)

    experimentDateAndTime = create_experiment_date_and_time_dict(experimentDateAndTimeCells)
    procedure_details = create_procedure_details_dict(procedureCells, read_dict)
    wellsInfo = create_wells_dict(wellCells, wellInfoCells)

    for key, value in data_ids.items():
        read_single(file, first_col_id, last_col_id, key, value, data_sheet, data_dict, time_dict, temperature_dict,
                    read_dict)

    settings = []
    settings = generate_nested_list(settings, settingsCells)
    settings = append_to_nested_list(settings, settingsCells)
    settings = append_single_value_to_nested_list(settings, str(equipment))
    settings = append_single_value_to_nested_list(settings, str(filePaths))
    settings = append_single_value_to_nested_list(settings, str(experimentDateAndTime))
    settings = append_single_value_to_nested_list(settings, str(procedure_details))
    settings = append_single_value_to_nested_list(settings, str(wellsInfo))

    formatted_cells = []
    formatted_cells = generate_single_nested_list(formatted_cells)
    formatted_cells = append_single_value_to_nested_list(formatted_cells, identifier)
    formatted_cells = append_cells_to_single_nested_list(formatted_cells, str(time_dict))
    formatted_cells = append_single_value_to_nested_list(formatted_cells, str(temperature_dict))
    formatted_cells = append_cells_to_single_nested_list(formatted_cells, str(data_dict))

    add_to_database(formatted_cells)
    add_to_settings(settings)


def read_single(file, first_col_id, last_col_id, key, value, data_sheet, data_dict, time_dict, temperature_dict,
                read_dict):
    first_row_id = value[0]
    last_row_id = value[1]
    data_cells_id = (get_cellId(offset_col(first_col_id, 2), first_row_id), get_cellId(last_col_id, last_row_id))
    time_cells_id = (get_cellId(first_col_id, first_row_id), get_cellId(first_col_id, last_row_id))
    temp_cells_id = (
        get_cellId(offset_col(first_col_id, 1), first_row_id), get_cellId(offset_col(first_col_id, 1), first_row_id))

    data_cells = get_cells(file, data_cells_id, data_sheet)
    time_cells = get_cells(file, time_cells_id, data_sheet)
    temperature_cells = get_single_value(temp_cells_id, file, data_sheet)

    data = []
    data = generate_nested_list(data, data_cells)
    data = append_to_nested_list(data, data_cells)
    data_dict[key] = data

    time = []
    time = append_time_to_list(time, time_cells)
    time_dict[key] = time

    temperature_dict[key] = temperature_cells

    read_dict = add_to_read_dict(file, data_sheet, read_dict, key, first_row_id)


def offset_col(char, offset):
    return str(chr(ord(char) + offset))


def get_cellId(char, lastRow):
    return str(char + lastRow)


def get_identifier(identifierId, file, sheet):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet]

    cell = sheet[identifierId[0]:identifierId[1]]
    return cell[0][0].value


def create_file_paths_dict(filePathsCells):
    file_paths_dict = create_dict("Experiment", filePathsCells[0][0].value)
    return add_to_dict(file_paths_dict, "Protocol", filePathsCells[1][0].value)


def create_experiment_date_and_time_dict(experimentDateAndTimeCells):
    experiment_date_and_time_dict = create_dict("Date", experimentDateAndTimeCells[0][0].value.strftime("%d/%m/%Y"))
    return add_to_dict(experiment_date_and_time_dict, "Time",
                       experimentDateAndTimeCells[1][0].value.strftime("%H:%M:%S"))


def create_temperature_dict(odTemperatureCells, gfpTemperatureCells, rfpTemperatureCells):
    temperature_dict = create_dict("OD", odTemperatureCells)
    temperature_dict = add_to_dict(temperature_dict, "GFP", gfpTemperatureCells)
    return add_to_dict(temperature_dict, "RFP", rfpTemperatureCells)


def create_time_dict(odTimeCells, gfpTimeCells, rfpTimeCells):
    odTime = []
    odTime = append_time_to_list(odTime, odTimeCells)
    gfpTime = []
    gfpTime = append_time_to_list(gfpTime, gfpTimeCells)
    rfpTime = []
    rfpTime = generate_nested_list(rfpTime, rfpTimeCells)
    rfpTime = append_time_to_list(rfpTime, rfpTimeCells)
    time_dict = create_dict("OD", odTime)
    time_dict = add_to_dict(time_dict, "GFP", gfpTime)
    return add_to_dict(time_dict, "RFP", rfpTime)


def create_data_dict(odCells, gfpCells, rfpCells):
    odValues = []
    odValues = generate_nested_list(odValues, odCells)
    odValues = append_to_nested_list(odValues, odCells)
    gfpValues = []
    gfpValues = generate_nested_list(gfpValues, gfpCells)
    gfpValues = append_to_nested_list(gfpValues, gfpCells)
    rfpValues = []
    rfpValues = generate_nested_list(rfpValues, rfpCells)
    rfpValues = append_to_nested_list(rfpValues, rfpCells)
    data_dict = create_dict("OD", odValues)
    data_dict = add_to_dict(data_dict, "GFP", gfpValues)
    return add_to_dict(data_dict, "RFP", rfpValues)


def create_read_info_dict(readODWavelength, readGFPCells, readRFPCells):
    read_info_dict = create_dict("OD", create_dict("Wavelength", extract_int_as_string(readODWavelength)))
    gfp_dict = create_fp_dict(readGFPCells)
    read_info_dict = add_to_dict(read_info_dict, "GFP", str(gfp_dict))
    rfp_dict = create_fp_dict(readRFPCells)
    read_info_dict = add_to_dict(read_info_dict, "RFP", str(rfp_dict))
    return read_info_dict


def add_to_read_dict(file, data_sheet, read_dict, key, first_row_id):
    read_id = ("A1", "B" + first_row_id)
    read_cells = get_cells(file, read_id, data_sheet)
    isRightExpType = False
    isCompleted = False

    for i in range(len(read_cells)):
        for j in range(len(read_cells[0])):
            isRightExpType = is_right_exp_type(isRightExpType, read_cells, i, j, key)
            if (isRightExpType and not isCompleted):
                if (is_add_read_cell(read_cells, i, j, key)):
                    add_info_to_read_dict(read_dict, read_cells, i, j, key)
                    isCompleted = True
    return read_dict

def add_info_to_read_dict(read_dict, read_cells, i, j, key):
    if key == "OD" and "Wavelength" in read_cells[i][j].value:
        info_dict = {}
        info_dict["Wavelength"] = extract_int_as_string(read_cells[i][j].value)
        read_dict[key] = info_dict
    if key =="GFP" or key == "RFP" and "Excitation" in read_cells[i][j].value:
        excitation_emissions = read_cells[i][j].value.split(",")
        info_dict = {}
        info_dict["Excitation"] = extract_int_as_string(excitation_emissions[0])
        info_dict["Emission"] = extract_int_as_string(excitation_emissions[1])
        info_dict["Gain"] = extract_int_as_string(read_cells[i+1][j].value)
        read_dict[key] = info_dict


def is_add_read_cell(read_cells, i, j, key):
    if read_cells[i][j].value is None:
        return False
    if not isinstance(read_cells[i][j].value, str):
        return False
    if key == "OD" and "Wavelength" in read_cells[i][j].value:
        return True
    if (key =="GFP" or key == "RFP") and "Excitation" in read_cells[i][j].value:
        return True

    return False


def is_right_exp_type(isRightExpType, read_cells, i, j, key):
    if (isRightExpType):
        return isRightExpType

    if (str(read_cells[i][j].value).lower() == "read" and str(read_cells[i][j + 1].value) == key):
        # print(read_cells[i][j].coordinate)
        return True

    return False


def create_procedure_details_dict(procedureCells, readInfo):
    procedure_details_dict = create_dict("Plate Type", procedureCells[0][0].value)
    procedure_details_dict = add_to_dict(procedure_details_dict, "Set Temperature", procedureCells[2][0].value)
    procedure_details_dict = add_to_dict(procedure_details_dict, "Start Kinetic", procedureCells[5][0].value)
    shake_dict = create_shake_dict(procedureCells)
    procedure_details_dict = add_to_dict(procedure_details_dict, "Shake", str(shake_dict))

    return add_to_dict(procedure_details_dict, "Read", readInfo)


def create_wells_dict(wellCells, wellInfoCells):
    wells_dict = {}
    for i in range(len(wellCells[0])):
        wells_dict = add_to_dict(wells_dict, wellCells[0][i].value, wellInfoCells[0][i].value)

    return wells_dict


def create_shake_dict(procedureCells):
    shake_dict = create_dict("Orbital", procedureCells[6][0].value.replace("Orbital: ", ""))
    return add_to_dict(shake_dict, "Frequency", procedureCells[7][0].value.replace("Frequency: ", ""))


def create_fp_dict(readFPCells):
    excitation_emissions = readFPCells[0][0].value.split(",")
    gfp_dict = create_dict("Excitation", extract_int_as_string(excitation_emissions[0]))
    gfp_dict = add_to_dict(gfp_dict, "Emission", extract_int_as_string(excitation_emissions[1]))
    return add_to_dict(gfp_dict, "Gain", extract_int_as_string(readFPCells[1][0].value))


def get_cells(file, cellsId, sheet):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet]

    cells = sheet[cellsId[0]: cellsId[1]]

    return cells


def extract_int_as_string(string):
    return str(re.search(r'\d+', string).group())


def get_single_value(temperatureId, file, sheet):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet]

    cell = sheet[temperatureId[0]:temperatureId[1]]
    return cell[0][0].value


def create_dict(key, value):
    output_dict = {key: value}
    return output_dict


def add_to_dict(dict, key, value):
    dict[key] = value
    return dict


def generate_nested_list(formatted_cells, cells):
    for i in range(len(cells)):
        formatted_cells.append([])

    return formatted_cells


def generate_single_nested_list(formatted_cells):
    formatted_cells.append([])

    return formatted_cells


def append_to_nested_list(formatted_cells, cells):
    for i in range(len(cells)):
        for k in range(len(cells[0])):
            formatted_cells[i].append(cells[i][k].value)

    return formatted_cells


def append_single_value_to_nested_list(formatted_cells, value):
    formatted_cells[0].append(value)

    return formatted_cells


def append_cells_to_single_nested_list(formatted_cells, cells):
    to_append = ''
    for i in range(len(cells)):
        to_append += str(cells[i])
    formatted_cells[0].append(to_append)
    return formatted_cells


def append_values_to_nested_list(formatted_cells, cells):
    for i in range(len(cells)):
        formatted_cells[i].append(str(cells[i]))

    return formatted_cells


def append_time_to_nested_list(formatted_cells, cells):
    for i in range(len(cells)):
        for k in range(len(cells[0])):
            formatted_cells[i].append(cells[i][k].value.strftime("%H:%M:%S"))

    return formatted_cells


def append_time_to_list(formatted_cells, cells):
    for i in range(len(cells)):
        formatted_cells.append(cells[i][0].value.strftime("%H:%M:%S"))

    return formatted_cells


def append_identifier_to_nested_list(formatted_cells, cells, identifier):
    for i in range(len(cells)):
        formatted_cells[i].append(identifier)

    return formatted_cells


def append_temperature_to_nested_list(formatted_cells, cells, temperature):
    for i in range(len(cells)):
        formatted_cells[i].append(temperature)

    return formatted_cells


def add_to_database(cells):
    db = "sqlite:///EBase.db"
    create_table(db)
    connection = create_connection("EBase.db")

    connection.executemany("""

                           INSERT INTO
                           data(experimental_id, time, set_temperature, data_values)
                           VALUES(?,?,?,?)""", cells)

    connection.commit()
    connection.close()


def add_to_settings(cells):
    db = "sqlite:///EBase.db"
    create_table(db)
    connection = create_connection("EBase.db")

    connection.executemany("""

                           INSERT INTO
                           settings(experimental_id, measurement, experiment_type, temperature, media, plasmid_name, 
                           equipment, filepath, date, procedure_details, wells_info )
                           VALUES(?,?,?,?,?,?,?,?,?,?,?)""", cells)

    connection.commit()
    connection.close()


def create_connection(db_file):
    """ create a database connection to the SQLite database
        specified by db_file
    :param db_file: database file
    :return: Connection object or None
    """
    conn = None
    try:
        conn = sqlite3.connect(db_file)
        return conn
    except sqlite3.Error as e:
        print(e)

    return conn


def create_table(db):
    engine = create_engine(db)

    meta = MetaData()

    data = Table(
        'data', meta,
        Column('experimental_id', String, primary_key=True),
        Column('time', String),
        Column('set_temperature', String),
        Column('data_values', String)
    )

    settings = Table(
        'settings', meta,
        Column('experimental_id', String, primary_key=True),
        Column('measurement', String),
        Column('experiment_type', String),
        Column('Temperature', String),
        Column('Media', String),
        Column('plasmid_name', String),
        Column('Equipment', String),
        Column('Filepath', String),
        Column('Date', String),
        Column('procedure_details', String),
        Column('wells_info', String)
    )
    meta.create_all(engine)


# Enable the script to be run from the command line
if __name__ == "__main__":
    data_id = {}
    ids = ["63", "95"]
    data_id["OD"] = ids
    data_id["GFP"] = ["100", "132"]
    data_id["RFP"] = ["137", "169"]

    read("cytation_H1_plate1.xlsx", "B", "CU", data_id, "Microplate reader")
